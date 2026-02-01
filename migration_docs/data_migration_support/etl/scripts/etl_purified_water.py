"""
ETL: Purified Water Test Reports (정제수성적서) -> Supabase

Requirements:
- pandas, openpyxl, supabase-py (minimal deps)
- Supabase service role key via SUPABASE_SERVICE_ROLE_KEY
"""

from __future__ import annotations

import os
import re
from datetime import date
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook
from supabase import Client, create_client

try:
    import xlrd

    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

PROJECT_ROOT = Path(__file__).resolve().parents[4]


def load_env_file(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line or line.strip().startswith("#"):
            continue
        if "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ.setdefault(key.strip(), value.strip())


load_env_file(PROJECT_ROOT / ".env")

SUPABASE_URL = os.getenv("SUPABASE_URL", "https://usvjbuudnofwhmclwhfl.supabase.co")
SUPABASE_SERVICE_ROLE_KEY = os.getenv("SUPABASE_SERVICE_ROLE_KEY", "")

SOURCE_DIR = Path("migration_docs/서식 샘플/정제수")

TESTS_TABLE = "labdoc_demo_purified_water_tests"
RESULTS_TABLE = "labdoc_demo_purified_water_test_results"

TEST_ITEMS = [
    {
        "row": 12,
        "code": "appearance",
        "name": "성상",
        "spec": "무색투명액상,무취,무미",
        "freq": "daily",
    },
    {"row": 13, "code": "ph", "name": "pH", "spec": "5.0~7.0", "freq": "daily"},
    {
        "row": 14,
        "code": "chloride",
        "name": "염화물",
        "spec": "액이 변해선 안됨",
        "freq": "daily",
    },
    {
        "row": 15,
        "code": "sulfate",
        "name": "황산염",
        "spec": "액이 변해선 안됨",
        "freq": "daily",
    },
    {
        "row": 16,
        "code": "heavy_metals",
        "name": "중금속",
        "spec": "비교액보다 진해선 안됨",
        "freq": "daily",
    },
    {
        "row": 17,
        "code": "residual_chlorine",
        "name": "잔류염소",
        "spec": "색을 나타내선 안됨",
        "freq": "weekly",
    },
    {
        "row": 18,
        "code": "ammonia",
        "name": "암모니아",
        "spec": "액이 변해선 안됨",
        "freq": "weekly",
    },
    {
        "row": 19,
        "code": "carbon_dioxide",
        "name": "이산화탄소",
        "spec": "액이 변해선 안됨",
        "freq": "weekly",
    },
    {
        "row": 20,
        "code": "potassium",
        "name": "칼륨",
        "spec": "액이 변해선 안됨",
        "freq": "weekly",
    },
    {
        "row": 21,
        "code": "permanganate",
        "name": "과망간산칼륨환원성물질",
        "spec": "홍색이 없어져선 안됨",
        "freq": "weekly",
    },
    {
        "row": 22,
        "code": "evaporation_residue",
        "name": "증발잔류물",
        "spec": "1mg 이하",
        "freq": "weekly",
    },
    {
        "row": 23,
        "code": "microorganism",
        "name": "미생물",
        "spec": "100 CFU/g 이하",
        "freq": "weekly",
    },
]

RESULT_COL = "H"
JUDGMENT_COL = "J"


def parse_filename(filename: str) -> tuple[int, str]:
    match = re.match(r"정제수성적서_(\d{4})년\s*(상반기|하반기)", filename)
    if match:
        return int(match.group(1)), match.group(2)

    match = re.match(r"정제수성적서\((\d{4})년\s*후반기\)", filename)
    if match:
        return int(match.group(1)), "하반기"

    match = re.match(r"정제수성적서\((\d{4})년(\d{1,2})월\)", filename)
    if match:
        year = int(match.group(1))
        month = int(match.group(2))
        half = "상반기" if month <= 6 else "하반기"
        return year, half

    raise ValueError(f"Unknown filename pattern: {filename}")


def parse_sheet_name(sheet_name: str, year: int, half: str) -> date | None:
    skip_patterns = ["Sheet", "총괄", "목차", "요약"]
    if any(p in sheet_name for p in skip_patterns):
        return None

    match = re.match(r"(\d{1,2})월\s*(\d{1,2})일", sheet_name)
    if not match:
        return None

    month = int(match.group(1))
    day = int(match.group(2))

    if half == "상반기" and month > 6:
        return None
    if half == "하반기" and month <= 6:
        year += 1

    try:
        return date(year, month, day)
    except ValueError:
        return None


def parse_test_date_from_cell(cell_value: Any, sheet_date: date | None) -> date | None:
    if not cell_value:
        return sheet_date

    match = re.search(
        r"(\d{4})\s*년\s*(\d{1,2})\s*월\s*(\d{1,2})\s*일",
        str(cell_value),
    )
    if match:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))

    return sheet_date


def extract_numeric_result(item_code: str, result_value: Any) -> float | None:
    if item_code != "ph" or result_value is None:
        return None
    try:
        value = float(str(result_value).strip())
        return value if 0 <= value <= 14 else None
    except (TypeError, ValueError):
        return None


def normalize_result_value(raw_value: Any) -> str | None:
    if raw_value is None:
        return None
    val = str(raw_value).strip()
    if val == "" or val in {"해당사항 없음", "해당사항없음"}:
        return None
    return val


def convert_judgment(
    raw_value: Any,
    item_code: str,
    result_value: str | None,
    prev_judgment: str | None = None,
) -> str:
    if raw_value is None or str(raw_value).strip() == "":
        return "NA"

    val = str(raw_value).strip()
    mapping = {
        "적합": "PASS",
        "부적합": "FAIL",
        "해당사항 없음": "NA",
        "해당사항없음": "NA",
        "불검출": "PASS",
    }
    if val in mapping:
        return mapping[val]

    if val in ["〃", '"', "''", "〃〃"]:
        return prev_judgment if prev_judgment else "NA"

    if item_code == "ph" and result_value:
        try:
            ph = float(result_value)
            return "PASS" if 5.0 <= ph <= 7.0 else "FAIL"
        except (TypeError, ValueError):
            return "NA"

    return "NA"


def calculate_overall_result(results: Iterable[dict[str, Any]]) -> str:
    for result in results:
        if result.get("judgment") == "FAIL":
            return "부적합"
    return "적합"


def get_cell_value(sheet: Any, row: int, col: str) -> Any:
    if hasattr(sheet, "cell_value"):
        col_idx = ord(col.upper()) - ord("A")
        return sheet.cell_value(row - 1, col_idx)
    return sheet[f"{col}{row}"].value


def should_skip_file(filename: str) -> bool:
    return filename in {"정제수pH통계.xls", "정제수성적서.xls"}


def collect_source_files() -> list[Path]:
    files: list[Path] = []
    for pattern in ("정제수성적서*.xlsx", "정제수성적서*.xls"):
        for path in SOURCE_DIR.glob(pattern):
            if should_skip_file(path.name):
                continue
            files.append(path)
    return sorted(files)


def load_workbook_compat(file_path: Path) -> tuple[Any, list[str]]:
    file_ext = file_path.suffix.lower()
    if file_ext == ".xls":
        if not HAS_XLRD:
            raise RuntimeError("xlrd is required to read .xls files")
        workbook = xlrd.open_workbook(file_path)
        return workbook, workbook.sheet_names()
    workbook = load_workbook(file_path, data_only=True, read_only=True)
    return workbook, workbook.sheetnames


def get_sheet(workbook: Any, sheet_name: str) -> Any:
    if hasattr(workbook, "sheet_by_name"):
        return workbook.sheet_by_name(sheet_name)
    return workbook[sheet_name]


def build_results(sheet) -> list[dict[str, Any]]:
    results: list[dict[str, Any]] = []
    prev_judgment: str | None = None

    for item in TEST_ITEMS:
        raw_value = get_cell_value(sheet, item["row"], RESULT_COL)
        raw_judgment = get_cell_value(sheet, item["row"], JUDGMENT_COL)

        result_value = normalize_result_value(raw_value)
        judgment = convert_judgment(
            raw_judgment, item["code"], result_value, prev_judgment
        )

        if result_value is None and item["freq"] == "weekly":
            judgment = "NA"

        result_numeric = extract_numeric_result(item["code"], result_value)
        if judgment != "NA":
            prev_judgment = judgment

        results.append(
            {
                "test_item_code": item["code"],
                "test_item_name": item["name"],
                "specification": item["spec"],
                "result_value": result_value,
                "result_numeric": result_numeric,
                "judgment": judgment,
                "test_frequency": item["freq"],
            }
        )

    return results


def upsert_test(client: Client, payload: dict[str, Any]) -> str:
    client.table(TESTS_TABLE).upsert(
        payload,
        on_conflict="test_date,source_file",
    ).execute()

    response = (
        client.table(TESTS_TABLE)
        .select("id")
        .eq("test_date", payload["test_date"])
        .eq("source_file", payload["source_file"])
        .limit(1)
        .execute()
    )
    if not response.data:
        raise RuntimeError("Failed to fetch test id after upsert")
    return response.data[0]["id"]


def upsert_results(client: Client, results_payload: list[dict[str, Any]]) -> None:
    if not results_payload:
        return
    client.table(RESULTS_TABLE).upsert(
        results_payload,
        on_conflict="test_id,test_item_code",
    ).execute()


def main() -> None:
    print("=" * 70)
    print("Purified Water ETL")
    print("=" * 70)

    if not SUPABASE_SERVICE_ROLE_KEY:
        raise RuntimeError("SUPABASE_SERVICE_ROLE_KEY is required")

    client = create_client(SUPABASE_URL, SUPABASE_SERVICE_ROLE_KEY)
    source_files = collect_source_files()

    if not source_files:
        print("No source files found.")
        return

    total_tests = 0
    total_results = 0
    errors: list[str] = []

    for file_path in source_files:
        print(f"\n[FILE] {file_path.name}")
        try:
            year, half = parse_filename(file_path.stem)
        except ValueError as exc:
            errors.append(f"{file_path.name}: {exc}")
            continue

        try:
            workbook, sheet_names = load_workbook_compat(file_path)
        except Exception as exc:
            errors.append(f"{file_path.name}: failed to load workbook ({exc})")
            continue

        for sheet_name in sheet_names:
            sheet_date = parse_sheet_name(sheet_name, year, half)
            if not sheet_date:
                continue

            print(f"  [SHEET] {sheet_name}")
            try:
                sheet = get_sheet(workbook, sheet_name)

                test_date = parse_test_date_from_cell(
                    get_cell_value(sheet, 5, "A"), sheet_date
                )
                if not test_date:
                    raise RuntimeError("test_date not found")

                header = {
                    "test_date": test_date.isoformat(),
                    "material_name": get_cell_value(sheet, 7, "C"),
                    "sample_amount": get_cell_value(sheet, 7, "D"),
                    "sampling_location": get_cell_value(sheet, 7, "I"),
                    "collector": get_cell_value(sheet, 7, "K"),
                    "inspector": get_cell_value(sheet, 25, "I"),
                    "source_file": file_path.name,
                    "source_sheet": sheet_name,
                    "source_row": None,
                }

                results = build_results(sheet)
                header["overall_result"] = calculate_overall_result(results)

                test_id = upsert_test(client, header)

                results_payload = [{"test_id": test_id, **result} for result in results]
                upsert_results(client, results_payload)

                total_tests += 1
                total_results += len(results_payload)
            except Exception as exc:
                errors.append(f"{file_path.name}::{sheet_name}: {exc}")
                continue

    print("\n" + "=" * 70)
    print(f"TOTAL TESTS: {total_tests}")
    print(f"TOTAL RESULTS: {total_results}")
    if errors:
        print(f"ERRORS ({len(errors)}):")
        for err in errors:
            print(f"  - {err}")
    print("=" * 70)


if __name__ == "__main__":
    main()
